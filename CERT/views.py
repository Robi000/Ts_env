from django.shortcuts import render, redirect, get_object_or_404
from .models import Student, Payment_log, CourceAndPrice
from .forms import PaymentStatusForm
from django.db.models import Max
from django.db.models import Q
import urllib.parse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime, timedelta
from reportlab.lib.utils import ImageReader
from io import BytesIO
from .utils import download_image
from django.http import HttpResponse
import qrcode
import pandas as pd
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook

@login_required
def Payment(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Registeral" or request.user.users.role == "Sub-Admin":
        max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
        lis = list(range(max_value + 1)) if max_value else []
        liss = lis[1:]
        liss.reverse()
        context = {"Batch": liss}
        context["Batc"] = max_value

        if request.htmx:
            print(request.POST)
            Students = Student.objects.filter(Batch=int(request.POST["Batch"])).order_by(
                'cource', 'first_name', 'last_name')
            context["stus"] = Students
            context["Batch"] = int(request.POST["Batch"])
            return render(request, "p/payment.html", context)
        Students = Student.objects.filter(Batch=max_value).order_by(
            'cource', 'first_name', 'last_name')
        context["stus"] = Students
        return render(request, "payment.html", context)
    else:
        return render(request, 'Unautorized.html', {})


def payment_handler(request):

    context = {}
    print(request.POST)
    print(request.POST["Batc"])
    batc = request.POST["Batc"]
    Students = Student.objects.filter(Batch=int(request.POST["Batc"]))
    count = 0
    for stu in Students:
        if stu.payment_status != request.POST.get(str(stu.id), stu.payment_status):
            stu.payment_status = request.POST.get(
                str(stu.id), stu.payment_status)
            print(stu.cource)
            amount = CourceAndPrice.objects.get(name=stu.cource)
            amount = amount.price
            filterd = Payment_log.objects.filter(student=stu)
            if filterd:
                for x in filterd:
                    x.delete()
                    Payment_log.objects.create(
                        user=request.user.users, amount=amount, student=stu, field=stu.cource)
            else:
                Payment_log.objects.create(
                    user=request.user.users, amount=amount, student=stu, field=stu.cource)
        stu.save()
        count += 1
    context["stus"] = Students
    context["Batch"] = int(request.POST["Batc"])
    print(int(request.POST["Batc"]))
    if count > 0:
        context["message"] = "scussessfully changed"
    context["id"] = True

    return render(request, "p/payment.html", context)


@login_required
def complition(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Teacher" or request.user.users.role == "Sub-Admin":

        max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
        lis = list(range(max_value + 1))
        liss = lis[1:]
        liss.reverse()

        context = {"Batch": max_value}
        context["bbt"] = liss

        if request.htmx:
            print(request.POST)
            Students = Student.objects.filter(Batch=int(request.POST["Batch"])).order_by(
                "finished", 'cource', 'first_name', 'last_name')
            context["stus"] = Students
            context["Batch"] = int(request.POST["Batch"])
            return render(request, "p/complition.html", context)
        Students = Student.objects.filter(Batch=max_value).order_by(
            "finished", 'cource', 'first_name', 'last_name')
        context["stus"] = Students
        return render(request, "complition.html", context)
    else:
        return render(request, 'Unautorized.html', {})


def complition_handler(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Teacher" or request.user.users.role == "Sub-Admin":
        dic = {"on": True, "off": False}
        context = {}
        print(request.POST["Batch"])
        Students = Student.objects.filter(Batch=int(request.POST["Batch"]))
        count = 0
        for stu in Students:
            if request.POST.get(str(stu.id), False):
                if not stu.finished:
                    stu.finished = True
                    stu.save()
                    count += 1
            else:
                if stu.finished:
                    stu.finished = False
                    stu.save()
                    count += 1

        context["stus"] = Students.order_by(
            "finished", 'cource', 'first_name', 'last_name')
        context["Batch"] = int(request.POST["Batch"])
        if count > 0:
            context["message"] = "scussessfully changed"
        context["id"] = True
        return render(request, "p/complition.html", context)
    else:
        return render(request, 'Unautorized.html', {})


@login_required
def uncomplition(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Teacher" or request.user.users.role == "Sub-Admin":
        max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
        lis = list(range(max_value + 1)) if max_value else []
        liss = lis[1:]
        liss.reverse()
        context = {"Batch": liss}

        if request.htmx:
            print(request.POST)
            Students = Student.objects.filter(Batch=int(request.POST["Batch"])).order_by(
                'cource', 'first_name', 'last_name')
            context["stus"] = Students.filter(finished=False)
            context["Batch"] = int(request.POST["Batch"])
            return render(request, "p/uncomplition.html", context)
        Students = Student.objects.filter(Batch=max_value).order_by(
            'cource', 'first_name', 'last_name')
        context["stus"] = Students.filter(finished=False)
        return render(request, "uncomplition.html", context)
    else:
        return render(request, 'Unautorized.html', {})


def uncomplition_handler(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Teacher" or request.user.users.role == "Sub-Admin":
        dic = {"on": True, "off": False}
        context = {}
        Students = Student.objects.filter(Batch=int(request.POST["Batch"]))
        count = 0
        for stu in Students:
            if request.POST.get(str(stu.id), False):
                if not stu.finished:
                    stu.finished = True
                    stu.save()
                    count += 1

        context["stus"] = Students.filter(finished=False)
        context["Batch"] = int(request.POST["Batch"])
        if count > 0:
            context["message"] = "scussessfully changed"
        context["id"] = True
        return render(request, "p/uncomplition.html", context)
    else:
        return render(request, 'Unautorized.html', {})


@login_required
def refund(request):
    if request.user.users.role != 'Admin' or request.user.users.role == "Sub-Admin":
        return render(request, 'Unautorized.html', {})
    max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
    lis = list(range(max_value + 1)) if max_value else []
    liss = lis[1:]
    liss.reverse()
    context = {"Batch": liss}
    context["Batc"] = max_value

    if request.htmx:
        print(request.POST)
        Students = Student.objects.filter(Batch=int(request.POST["Batch"])).filter(
            payment_status="paid"
        ).order_by('cource', 'first_name', 'last_name')
        context["stus"] = Students
        context["Batch"] = int(request.POST["Batch"])
        context["Batc"] = int(request.POST["Batch"])
        return render(request, "p/refund.html", context)
    Students = Student.objects.filter(
        Batch=max_value).filter(payment_status="paid").order_by('cource', 'first_name', 'last_name')
    context["stus"] = Students

    return render(request, "refund.html", context)


def refund_handler(request):
    if request.user.users.role != 'Admin' or request.user.users.role == "Sub-Admin":
        return render(request, 'Unautorized.html', {})
    context = {}
    print(request.POST)
    Students = Student.objects.filter(Batch=int(request.POST["Batc"])).filter(
        payment_status="paid"
    )
    count = 0
    for stu in Students:
        print(stu.id)
        if stu.payment_status != request.POST.get(str(stu.id), stu.payment_status):
            stu.payment_status = request.POST.get(
                str(stu.id), stu.payment_status)
            print("changed")
            paymentlog = Payment_log.objects.filter(student=stu)
            if len(paymentlog) == 1:
                print("if here ......")
                paymentlog[0].refunded = True
                paymentlog[0].save()
                print(paymentlog[0].updated)
            else:
                print("else here ...... ")
                for x in paymentlog:
                    x.delete()
                amount = CourceAndPrice.objects.get(name=stu.cource)
                amount = amount.price
                Payment_log.objects.create(
                    user=request.user.users, amount=amount, student=stu, field=stu.cource, refunded=True)
            stu.save()
            count += 1
    context["stus"] = Students.filter(payment_status="paid")
    context["Batc"] = int(request.POST["Batc"])
    if count > 0:
        context["message"] = "scussessfully changed"
    context["id"] = True
    return render(request, "p/refund.html", context)


@login_required
def unpaid(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Registeral" or request.user.users.role == "Sub-Admin":

        max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
        lis = list(range(max_value + 1)) if max_value else []
        liss = lis[1:]
        liss.reverse()
        context = {"Batch": liss}

        if request.htmx:
            print(request.POST)
            Students = Student.objects.filter(Batch=int(request.POST["Batch"])).filter(
                payment_status="pending"
            ).order_by('cource', 'first_name', 'last_name')
            context["stus"] = Students
            context["Batch"] = int(request.POST["Batch"])
            return render(request, "p/unpaid.html", context)
        Students = Student.objects.filter(
            Batch=max_value).filter(payment_status="pending").order_by('cource', 'first_name', 'last_name')
        context["stus"] = Students
        return render(request, "unpaid.html", context)
    else:
        return render(request, 'Unautorized.html', {})


def unpaid_handler(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Registeral" or request.user.users.role == "Sub-Admin":
        context = {}
        Students = Student.objects.filter(Batch=int(request.POST["Batch"])).filter(
            payment_status="pending"
        )
        count = 0
        for stu in Students:
            print(stu.id)
            if stu.payment_status != request.POST.get(str(stu.id), stu.payment_status):
                stu.payment_status = request.POST.get(
                    str(stu.id), stu.payment_status)
                print("changed")
                stu.save()
                count += 1
        context["stus"] = Students.filter(payment_status="pending")
        context["Batch"] = int(request.POST["Batch"])
        if count > 0:
            context["message"] = "scussessfully changed"
        context["id"] = True
        return render(request, "p/unpaid.html", context)
    else:
        return render(request, 'Unautorized.html', {})


@login_required
def student_detail_search(request):

    context = {}
    if request.htmx:
        name = request.GET.get('name', '')
        students = Student.objects.filter(
            Q(first_name__icontains=name) | Q(last_name__icontains=name))
        if " " in name:
            name = name.split()
            if len(name) > 1:
                students = Student.objects.filter(
                Q(first_name__icontains=name[0]) | Q(last_name__icontains=name[1]) | Q(first_name__icontains=name[1]) | Q(last_name__icontains=name[0]) )
        
        context['stus'] = students.order_by(
            "-Batch" ,'cource', 'first_name', 'last_name')
        return render(request, "p/search_result.html", context)

    return render(request, "student_detail_search.html", context)


def extract_id_from_link(link):
    parsed_url = urllib.parse.urlparse(link)
    query_params = urllib.parse.parse_qs(parsed_url.query)

    if 'id' in query_params:
        return query_params['id'][0]
    else:
        return None


def student_detail(request, id):
    context = {}
    student = get_object_or_404(Student, id=id)
    context['stu'] = student

    link = 'https://drive.google.com/uc?export=view&id=' + \
        extract_id_from_link(student.photo_url)
    context["link"] = link

    return render(request, "student_detail.html", context)


def generate_pdf(student=()):

    image_path = 'CERT/img/certeficate.png'

    pdfmetrics.registerFont(TTFont('Impact', 'impact.ttf'))

    # Create the canvas
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=landscape(A4))
    for stu in student:
        image_url = 'https://drive.google.com/uc?export=view&id=' + \
            extract_id_from_link(stu.photo_url)

        # Draw the image on the canvas
        c.drawImage(image_path, -13, 0, width=865, height=600)

        # Your content drawing code here
        text = stu.first_name + " " + stu.last_name
        x, y = 630, 310
        c.setFont("Impact", 32)
        c.drawRightString(x, y, text)

        # Date of issue and date of expiry
        issue_date = datetime.now().strftime("%Y-%m-%d")
        stu.cert_rec_date = datetime.now()
        x, y = 270, 150
        c.setFont("Impact", 12)
        c.drawRightString(x, y, issue_date)

        current_date = datetime.now() + timedelta(days=365 * 2)
        expire_date = current_date.strftime("%Y-%m-%d")
        x, y = 390, 150
        c.setFont("Impact", 12)
        c.drawRightString(x, y, expire_date)

        # Type of certification
        c.setFont("Helvetica", 12)
        c.drawString(430, 255, stu.cource)

        # User photo
        try:
            print(str(stu.id)+stu.first_name)
            x = download_image(image_url, str(stu.id)+stu.first_name)
            print("downloded")
            c.drawImage(
                f'CERT/img/{str(stu.id)+stu.first_name}.{x}', 130, 350, width=120, height=150)
            print("writtem")
        except Exception as e:
            print(100*"*")
            print(f"Error: {e}")
            return False

        data_for_qr_code = f"https://samisafety.org/cert/detail/{stu.id}"
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data_for_qr_code)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        img.save(f"CERT/img/qr/{stu.id}.png")
        c.drawInlineImage(
            f"CERT/img/qr/{stu.id}.png", 680, 115, width=130, height=130)

        # End the current page
        c.showPage()

        stu.report_Generated = True
        stu.cert_rec_date = datetime.now()
        stu.save()

    # Save the PDF
    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer


@login_required
def batch_report(request):
    if request.user.users.role != 'Admin' or request.user.users.role == "Sub-Admin":
        return render(request, 'Unautorized.html', {})
    max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
    lis = list(range(max_value + 1))
    liss = lis[1:]
    liss.reverse()
    context = {"Batch": liss}
    if request.htmx:
        print(request.POST)
        Students = Student.objects.filter(Q(Batch=int(request.POST["Batch"])) & Q(
            payment_status='paid') & Q(finished=True) & Q(report_Generated=False))
        context["stus"] = Students.order_by(
            'finished', 'cource', 'first_name', 'last_name')
        context["current"] = int(request.POST["Batch"])
        return render(request, "p/report.html", context)
    Students = Student.objects.filter(Q(Batch=max_value) & Q(
        payment_status='paid') & Q(finished=True) & Q(report_Generated=False))
    context["current"] = max_value
    context["stus"] = Students.order_by('cource', 'first_name', 'last_name')
    return render(request, "report.html", context)


def batch_report_handler(request, batch):
    if request.user.users.role != 'Admin' or request.user.users.role == "Sub-Admin":
        return render(request, 'Unautorized.html', {})
    Students = Student.objects.filter(Q(Batch=batch) & Q(
        payment_status='paid') & Q(finished=True))
    pdf_buffer = generate_pdf(Students)
    if not pdf_buffer:
        return render(request, "gen_err.html", {})
    response = HttpResponse(pdf_buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Bulk_report_Batch_{batch}.pdf"'
    return response


def process_csv(file):
    # Your Pandas processing logic here
    df = pd.read_csv(file)
    # Process the dataframe as needed
    # Inspect the existing column names
    existing_columns = df.columns.tolist()

    # Define a mapping for the columns you want to rename
    column_mapping = {
        4: 'Phone',
        6: 'Edu_lev',
        7: 'curr_ocp',
        8: 'Cource',
        9: 'photo_link',
        11: 'Gender',
        12: 'Cert_type',
        13: 'Org',
    }

    # Rename the specified columns
    df.rename(columns={
        df.columns[4]: column_mapping[4],
        df.columns[6]: column_mapping[6],
        df.columns[7]: column_mapping[7],
        df.columns[8]: column_mapping[8],
        df.columns[9]: column_mapping[9],
        df.columns[11]: column_mapping[11],
        df.columns[12]: column_mapping[12],
        df.columns[13]: column_mapping[13],
    }, inplace=True)

    df = df.assign(Cource=df['Cource'].str.split(', ')).explode(
        'Cource').reset_index(drop=True)
    return df


@login_required
def student_reg(request):
    print(request.user.users.role)
    if request.user.users.role == 'Admin' or request.user.users.role == "Registeral" or request.user.users.role == "Sub-Admin":

        if request.method == 'POST':
            try:
                file = request.FILES['csvfile']
                dataframe = process_csv(file)
                print(dataframe)
                batch = Student.objects.aggregate(
                    Max("Batch"))["Batch__max"]
                batch = batch + 1 if batch else 1
                for index, row in dataframe.iterrows():
                    date_str = row['Timestamp']
                    formatted_date = datetime.strptime(
                        date_str, '%d/%m/%Y %H:%M:%S')
                    stu_data = {
                        'registration': formatted_date,
                        'first_name': row['First name '],
                        'last_name': row['Last name'],
                        'Batch': batch,
                        'email': row['Email Address'],
                        'phone': int(row['Phone']),
                        'age': int(row['Age']),
                        'edu_lev': row['Edu_lev'],
                        'org': row['Org'],
                        'occ': row['curr_ocp'],
                        'cource': row['Cource'],
                        'cert_type': row['Cert_type'],
                        'photo_url': row['photo_link'],
                    }
                    student = Student.objects.create(**stu_data)
                    print(student.first_name, "registerd ")
                stus = Student.objects.filter(Batch=batch).order_by(
                    'cource', 'first_name', 'last_name')
                context = {
                    "stus": stus
                }
                context["BCH"] = batch
                # Do something with the dataframe (e.g., save to database, perform calculations)
                return render(request, 'p/result.html', context)
            except Exception as e:
                return render(request, 'p/err_result.html', {"e":e})

        return render(request, 'upload_csv.html')
    else:
        return render(request, 'Unautorized.html', {})


def Student_edit_search(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Registeral" or request.user.users.role == "Sub-Admin":
        context = {}
        if request.htmx:
            name = request.GET.get('name', '')
            students = Student.objects.filter(
                Q(first_name__icontains=name) | Q(last_name__icontains=name))
            if " " in name:
                name = name.split()
                if len(name) > 1:
                    students = Student.objects.filter(
                    Q(first_name__icontains=name[0]) | Q(last_name__icontains=name[1]) | Q(first_name__icontains=name[1]) | Q(last_name__icontains=name[0]) )
            context['stus'] = students.order_by(
                "-Batch",'cource', 'first_name', 'last_name')
            return render(request, "p/search_result_edit.html", context)

        return render(request, "student_edit_search.html", context)
    else:
        return render(request, 'Unautorized.html', {})


def student_edit_handler(request, id):
    if request.user.users.role == 'Admin' or request.user.users.role == "Registeral" or request.user.users.role == "Sub-Admin":
        student = Student.objects.get(id=id)
        context = {'stu': student}
        if request.method == "POST":
            student.first_name = request.POST['first_name']
            student.last_name = request.POST['last_name']
            student.email = request.POST['email']
            student.age = request.POST['age']
            student.phone = request.POST['phone']
            student.edu_lev = request.POST['edu_lev']
            student.org = request.POST['org']
            student.edu_lev = request.POST['edu_lev']
            student.occ = request.POST['occ']
            student.cource = request.POST['cource']
            student.save()
            return redirect('stu_detail', id=student.idv)
        return render(request, "student_edit.html", context)
    else:
        return render(request, 'Unautorized.html', {})


def single_student_report(request):
    if request.user.users.role == 'Admin' or request.user.users.role == "Sub-Admin":
        max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
        lis = list(range(max_value + 1))
        liss = lis[1:]
        liss.reverse()
        context = {"Batch": liss}

        if request.htmx:
            print(request.POST)
            Students = Student.objects.filter(
                Q(Batch=int(request.POST["Batch"])) & Q(finished=True)).order_by('cource', 'first_name', 'last_name')
            context["stus"] = Students
            context["Batch"] = int(request.POST["Batch"])
            return render(request, "p/Single_stu_rport.html", context)
        Students = Student.objects.filter(
            Q(Batch=max_value) & Q(finished=True)).order_by('cource', 'first_name', 'last_name')
        context["stus"] = Students
        return render(request, "Single_stu_report.html", context)
    else:
        return render(request, 'Unautorized.html', {})


def single_student_report_handler(request, id):
    if request.user.users.role == 'Admin' or request.user.users.role == "Sub-Admin":
        Students = Student.objects.filter(Q(id=id) & Q(finished=True))
        if Students:
            pdf_buffer = generate_pdf(Students)
            if not pdf_buffer:
                return render(request, "gen_err.html", {})
            response = HttpResponse(
                pdf_buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="example.pdf"'
            return response
        else:
            return render(request, 'p/error.html', {})
    else:
        return render(request, 'Unautorized.html', {})


def generate_excel(request):
    max_value = Student.objects.aggregate(Max("Batch"))["Batch__max"]
    lis = list(range(max_value + 1)) if max_value else []
    liss = lis[1:]
    liss.reverse()
    context = {"Batch": liss}
    if request.method == "POST":
        batchs = list(map(int ,request.POST.getlist('btc')))
        students = Student.objects.filter(Batch__in=batchs)
        output = generate_student_excel(students)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Student info.xlsx'
        output.seek(0)
        response.write(output.getvalue())

        return response
       


    return render(request, 'generate_excel.html', context)



def generate_student_excel (Stu):
    # Fetch all records from the Payment_log model
    stu_log = Stu.order_by("Batch", "org" , "cource" ,"first_name" ,"last_name",)

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Logs"
    
    # Define headers
    headers = ["Student name", "Phone" , "Age" ,  'ORG' , 'occopation' ,"Cource" ,"Payment_status", "Cource_Status", "Certeficate_Given" ,"certeficate_issue" , "Certeficate_expire" , "batch"]

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Iterate over each record and insert data into the Excel sheet
    for row_num, stu in enumerate(stu_log, 2):
        ws.cell(row=row_num, column=1, value=stu.first_name + " " + stu.last_name)
        ws.cell(row=row_num, column=2, value=stu.phone)
        ws.cell(row=row_num, column=3, value=stu.age)
        ws.cell(row=row_num, column=4, value=stu.org)
        ws.cell(row=row_num, column=5, value=stu.occ)
        ws.cell(row=row_num, column=6, value=stu.cource)
        ws.cell(row=row_num, column=7, value=stu.payment_status)
        ws.cell(row=row_num, column=8, value=stu.finished)
        ws.cell(row=row_num, column=9, value=stu.report_Generated)
        ws.cell(row=row_num, column=10, value=stu.cert_rec_date)
        ws.cell(row=row_num, column=11, value=stu.cert_rec_date + timedelta(days=365 * 2) if stu.cert_rec_date else None)
        ws.cell(row=row_num, column=12, value=stu.Batch)

    # Save the workbook to a BytesIO object
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output