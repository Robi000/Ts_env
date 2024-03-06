from django.shortcuts import render, redirect, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from .models import users
from django.views.decorators.csrf import csrf_exempt
from CERT.models import Payment_log
from datetime import datetime
from reportlab.pdfgen import canvas
from django.db.models import Sum
from openpyxl import Workbook


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        # Authenticate user
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Log the user in
            login(request, user)
            return redirect(
                "auth_home"
            )  # Redirect to the home page or any other desired page
        else:
            # Display an error message or handle authentication failure
            return render(
                request, "login.html", {
                    "error_message": "Invalid username or password"}
            )

    return render(request, "login.html")


def logout_view(request):
    context = {}
    logout(request)
    return redirect("auth_home")


def manage_user(request):
    userss = users.objects.all()
    # print(userss[0].id, userss[0].user.id)
    context = {"userss": userss}
    if request.htmx:
        return render(request, "manage.html", context)
    return render(request, "manage.html", context)


@csrf_exempt
def delete_user(request, id):
    user = User.objects.get(id=id)
    user.delete()
    userss = users.objects.all()
    # print(userss[0].id, userss[0].user.id)
    context = {"userss": userss}
    if request.htmx:
        return render(request, "manage.html", context)
    return render(request, "manage.html", context)


def edit_user(request, id):
    user = User.objects.get(id=id)
    context = {"user": user.users}
    if request.method == "POST":
        userr = user.users
        userr.first_name = request.POST["first_name"]
        userr.Last_name = request.POST["Last_name"]
        userr.email = request.POST["email"]
        userr.phone = request.POST["phone"]
        print(request.POST)
        condition = False
        if userr.role != request.POST["role"]:
            if not request.POST["password"]:
                condition = True
            userr.role = request.POST["role"]
            user.username = request.POST["first_name"]+"_" + \
                request.POST["role"]+"_" + str(users.objects.all().count())
            user.save()
        if request.POST["password"]:
            print("hello world")
            context[
                "message"
            ] = "This notification serves to inform you that, henceforth, your password will no longer be displayed for security reasons.\
            It is imperative that you securely store and remember your password. In the event that you forget your password, kindly \
            reach out to your administrator, who will assist you in the password reset process.\
            It is important to note that, for security purposes, passwords are no longer visible, even to administrators.\
            <br>We are pleased to confirm that the edits to your profile have been successful. Welcome to our platform!"
            context["password"] = request.POST["password"]
            context["username"] = user.username
            user.set_password(request.POST["password"])
            user.save()
        if condition:
            context["message"] = f"Your new user name is {userr.user.username}"
        userr.save()
        user = User.objects.get(id=id)
        context["user"] = user.users
        return render(request, "user_detail.html", context)
    return render(request, "edit_user.html", context)


def register(request):
    context = {}
    if request.method == "POST":
        dic = dict(request.POST)
        del dic["csrfmiddlewaretoken"]
        for x in dic:
            dic[x] = dic[x][0]
        user = users.objects.create(**dic)
        context = {"user": user}
        context[
            "message"
        ] = "This notification serves to inform you that, henceforth, your password will no longer be displayed for security reasons.\
            It is imperative that you securely store and remember your password. In the event that you forget your password, kindly \
            reach out to your administrator, who will assist you in the password reset process.\
            It is important to note that, for security purposes, passwords are no longer visible, even to administrators.\
            we are pleased to confirm that your registration was successful. Welcome to our platform!"
        context["password"] = dic["password"]
        context["username"] = user.user.username
        return render(request, "user_detail.html", context)

    return render(request, "registration.html", context)


def user_detail(request, id):
    user = User.objects.get(id=id)
    context = {"user": user.users}
    return render(request, "user_detail.html", context)


def home(request):
    context = {
        "Payment_log": Payment_log.objects.all().order_by('-updated')[:10]}

    if request.method == "POST":
        print(request.POST)
        from_date_str = request.POST.get('fromDate')
        to_date_str = request.POST.get('toDate')
        typ = request.POST.get('type')
        print(from_date_str, to_date_str)
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

        filtered_payments = Payment_log.objects.filter(
            created__range=[from_date, to_date])
        
        if typ == "excel":
            output = generate_payment_excel(filtered_payments)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=payment_logs from {from_date_str} to {to_date_str}.xlsx'
            output.seek(0)
            response.write(output.getvalue())

            return response
        
        context["Payment_log"] = filtered_payments

        # Get distinct users
        distinct_users = filtered_payments.values_list(
            'user', flat=True).distinct()

        # Specify the local file path where you want to save the PDF
        # Change this to your desired file path
        pdf_file_path = 'payment_report.pdf'

        # Create PDF document
        pdf = canvas.Canvas(pdf_file_path)

        image_path = "CERT/img/Logo.jpg"
        pdf.drawImage(image_path, 30, 780, width=60, height=40)
        pdf.setFont('Helvetica-Bold', 16)
        pdf.drawString(110, 800,
                       f"Sami safty income Report from {from_date_str} to {to_date_str}")
        pdf.setFont('Helvetica', 12)
        # Set up initial y-coordinate for content
        y_coordinate = 700

        # Iterate over distinct users and calculate sum of amounts
        overall_total = 0
        overall_refunded = 0
        for user in distinct_users:
            user_payments = filtered_payments.filter(user=user)
            user_total = user_payments.aggregate(
                Sum('amount'))['amount__sum'] or 0
            user_refund = user_payments.filter(refunded=True)
            print(user_payments)
            print("*"*10)
            print(user_refund)
            user_refund_total = user_refund.aggregate(
                Sum('amount'))['amount__sum'] or 0

            # Display user total in PDF
            username = users.objects.get(id=user)
            name = username.first_name + " " + username.Last_name
            pdf.drawString(100, y_coordinate, f"User: {name}, Total Amount: {user_total}, refunded:{user_refund_total}")
            y_coordinate -= 20  # Adjust y-coordinate for next line
            overall_total += user_total
            overall_refunded += user_refund_total

        # Display overall total in PDF

        pdf.drawString(100, y_coordinate-20,
                       "*"*70)
        y_coordinate -= 20
        pdf.drawString(100, y_coordinate-20,
                       f"Overall  Gross Total Amount: {overall_total}")
        y_coordinate -= 20
        pdf.drawString(100, y_coordinate-20,
                       f"Overall  Refunded Total Amount: {overall_refunded}")
        y_coordinate -= 20
        pdf.drawString(100, y_coordinate-20,
                       f"Overall  Net Total Amount: {overall_total - overall_refunded}")
        userxs = request.user.users.first_name.capitalize(
        ) + " " + request.user.users.Last_name.capitalize()
        pdf.drawString(380, 100,
                       f"Created By {userxs}")
        pdf.drawString(380, 70, "signature _____________")

        # Save the PDF content
        pdf.save()

        # Use HttpResponse to return the PDF file
        with open(pdf_file_path, 'rb') as pdf_file:
            response = HttpResponse(
                pdf_file.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="payment_report.pdf"'
            return response
    else:
        # Non-Htmx request
        return render(request, "home2.html", context)


def generate_payment_excel(logs):
    # Fetch all records from the Payment_log model
    payment_logs = logs.order_by("student__org", "user__first_name" ,"student__first_name", "created", "refunded" )

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Logs"
    
    # Define headers
    headers = ['A-Name', 'S-Name', 'ORG' ,"Cource", 'Amount', 'Created', 'Refunded']

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Iterate over each record and insert data into the Excel sheet
    for row_num, log in enumerate(payment_logs, 2):
        ws.cell(row=row_num, column=1, value=log.user.first_name + " " + log.user.Last_name)
        ws.cell(row=row_num, column=2, value=log.student.first_name + " " + log.student.last_name)
        ws.cell(row=row_num, column=3, value=log.student.org)
        ws.cell(row=row_num, column=4, value=log.student.cource)
        ws.cell(row=row_num, column=5, value=log.amount)
        ws.cell(row=row_num, column=6, value=log.created)
        ws.cell(row=row_num, column=7, value=log.refunded)

    # Calculate total gross, refunded amount, and net gross
    total_gross = sum(log.amount for log in payment_logs)
    refunded_amount = sum(log.amount for log in payment_logs if log.refunded)
    net_gross = total_gross - refunded_amount

    # Append summary information to the Excel sheet
    ws.append(['Total Gross', total_gross])
    ws.append(['Refunded Amount', refunded_amount])
    ws.append(['Net Gross', net_gross])

    # Save the workbook to a BytesIO object
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payment_logs.xlsx'
    output.seek(0)
    response.write(output.getvalue())

    return response